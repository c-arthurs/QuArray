import sys
from PyQt5 import QtGui, uic
import qdarkgraystyle
from scripts import DABanalysis, Overlay
import sys
import os
from PyQt5.QtCore import *
from PyQt5.QtWidgets import QFileDialog
from PyQt5 import QtCore  # QtWidgets
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtWidgets import QApplication, QMainWindow
import qimage2ndarray
import random
import numpy as np
import matplotlib.image as mpimg
from skimage import color
from PyQt5.QtWidgets import QGraphicsScene
from openslide import OpenSlide
import qimage2ndarray
from PyQt5.QtGui import QPen, QBrush
import numpy as np
import sys
import os
from skimage import color, filters, measure, morphology
from openpyxl import load_workbook
import json


# setup the Graphics scene to detect clicks
class GraphicsScene(QGraphicsScene):
    def __init__(self, parent=None):
        QGraphicsScene.__init__(self, parent)
        self.coords = []
        self.parent = MyWindow

    def mousePressEvent(self, event):
        pen = QPen(QtCore.Qt.black)
        pen.setWidthF(10)  # border width
        brush = QBrush(QtCore.Qt.transparent)
        x = event.scenePos().x()
        y = event.scenePos().y()
        btn = event.button()
        if btn == 1:  # left click
            self.coords.append((x, y))
            self.addEllipse(x - 25, y - 25, 50, 50, pen, brush)
            # self.addRect(x-25, y-25, 100, 100, pen, brush)
            # self.drawPolygon(x-25, y-25, 100, 100, pen, brush)
        if btn == 2 and len(self.coords) >= 1:  # right click
            pen = QPen(QtCore.Qt.red)
            brush = QBrush(QtCore.Qt.red)
            self.addEllipse(self.coords[-1][0] - 25, self.coords[-1][1] - 25, 50, 50, pen, brush)
            self.coords = self.coords[:-1]
        print(btn)
        print(self.coords)

    def overlay_cores(self, core_diameter, scale_index, overview, cores):  # removed - centroid, image, cores
        pen = QPen(QtCore.Qt.black)
        pen.setWidthF(6)  # border width
        brush = QBrush(QtCore.Qt.transparent)
        self.centroid = [(y, x) for (x, y) in self.coords]
        diameter = core_diameter / scale_index
        a = 0
        for y, x in self.centroid:
            self.addRect((x - (diameter / 2)), (y - (diameter / 2)), diameter, diameter, pen, brush)
            text = self.addText(cores[a])  # label
            text.setPos(x, y)
            font = QtGui.QFont()
            font.setPointSize(80)
            text.setFont(font)
            a = a + 1


class MyWindow(QMainWindow):
    def __init__(self):
        super(MyWindow, self).__init__()
        uic.loadUi('./scripts/1.ui', self)
        self.excel_layout = True  # TODO add button for this
        self.core_diameter = 6000

        self.load_ndpi.clicked.connect(lambda: self.loadndpi())
        self.load_excel.clicked.connect(lambda: self.read_excel())
        self.overlay.clicked.connect(
            lambda: self.scene.overlay_cores(self.core_diameter, self.scale_index, self.overview, self.cores))
        self.export_2.clicked.connect(lambda: self.export_images())

        self.init_scene()
        self.show()

    def init_scene(self):
        self.scene = GraphicsScene(self)
        self.graphicsView.setScene(self.scene)
        self.pixmap = QtWidgets.QGraphicsPixmapItem()
        self.scene.addItem(self.pixmap)

    def show_info(self, text):
        self.metadata.setText(text)

    def loadndpi(self):
        self.init_scene()
        # self.path, _ = QFileDialog.getOpenFileName(parent=self, caption='Open file',
        #                                            directory="/Users/callum/Desktop")
        self.path = "/Users/callum/Desktop/ndpitest/MS2_AA37.ndpi"
        if self.path:
            self.output = os.path.splitext(self.path)[0] + '_split'
            if not os.path.exists(self.output):  # make output directory
                os.mkdir(self.output)

            self.name = os.path.split(self.output)[-1]

            self.load_ndpi.setStyleSheet("background-color: rgb(0,90,0)")
            print(self.path)
            self.image = OpenSlide(self.path)

            print(self.path + ' read to memory')
            print('    slide format = ' + str(OpenSlide.detect_format(self.path)))
            if str(OpenSlide.detect_format(self.path)) == "aperio":
                try:
                    print('    Magnification = ' + str(self.image.properties['openslide.objective-power']))  # TODO
                    print('    Date = ' + str(self.image.properties['aperio.Date']))
                    print('    dimensions = ' + str(self.image.dimensions))
                    print('    level_downsamples = ' + str(self.image.level_downsamples))
                except KeyError:
                    pass
            if str(OpenSlide.detect_format(self.path)) == "hamamatsu":
                try:
                    self.show_info(f"""Magnification = {str(self.image.properties['hamamatsu.SourceLens'])}
Date = {str(self.image.properties['tiff.DateTime'])}\ndimensions = {str(self.image.dimensions)}
level_downsamples = {str(self.image.level_downsamples)}""")
                    print('    Magnification = ' + str(self.image.properties['hamamatsu.SourceLens']))
                    print('    Date = ' + str(self.image.properties['tiff.DateTime']))
                    print('    dimensions = ' + str(self.image.dimensions))
                    print('    level_downsamples = ' + str(self.image.level_downsamples))
                    self.macro_image = self.image.associated_images['macro']
                except KeyError:
                    pass
            self.overview_level_width = 3000  # this is the target overview size - it will find the nearest level
            # self.excel_layout = excel_layout
            # self.user_select = user_select
            # self.select_batch = select_batch
            # if self.select_batch:
            #     self.user_select = True
            # self.centroid = []
            # self.core_diameter = core_diameter
            self.get_overview()

    def get_overview(self):
        self.width_height = [
            (int(self.image.properties[f'openslide.level[{i}].width']),
             int(self.image.properties[f'openslide.level[{i}].height'])) for i
            in range(int(self.image.properties['openslide.level-count']))]
        width = [self.width_height[i][0] for i in range(len(self.width_height))]
        self.lvl = np.where(width == self.find_nearest(width, self.overview_level_width))[0][0]
        self.scale_index = self.width_height[0][0] / self.width_height[self.lvl][0]
        print(f"nearest found at - {self.find_nearest(width, self.overview_level_width)}"
              f" corresponding to level {self.lvl}")
        self.overview = np.array(self.image.read_region(location=(0, 0), level=self.lvl,
                                                        size=self.width_height[self.lvl]))
        self.showimage()

    def find_nearest(self, array, value):
        # https://stackoverflow.com/questions/2566412/find-nearest-value-in-numpy-array
        array = np.asarray(array)
        idx = (np.abs(array - value)).argmin()
        return array[idx]

    def showimage(self):
        img = qimage2ndarray.array2qimage(self.overview, normalize=True)
        img = QtGui.QPixmap(img)
        self.pixmap.setPixmap(img)
        self.graphicsView.fitInView(self.graphicsView.sceneRect(), QtCore.Qt.KeepAspectRatio)

    def read_excel(self):
        # self.excelpath, _ = QFileDialog.getOpenFileName(parent=self, caption='Open file',
        #                                                 directory="/Users/callum/Desktop")
        self.excelpath = "/Users/callum/Desktop/ndpitest/MS2_AA37"
        if self.excelpath:
            self.load_excel.setStyleSheet("background-color: rgb(0,90,0)")
            excelname = os.path.splitext(self.excelpath)[0] + '.xlsx'
            wb = load_workbook(excelname)
            ws = wb.worksheets[0]
            cores = []
            values = []
            rowcount = []
            if not self.excel_layout:
                for row in ws.iter_rows():
                    for cell in row:
                        values.append(cell.value)
                        if cell.value == 1:
                            cores.append((str(chr(int(cell.row) + 64))) + str(int(ord(cell.column_letter)) - 64))
            else:
                cores = []
                for col in ws.iter_cols():
                    for cell in col:
                        values.append(cell.value)
                        if cell.value == 1:
                            cores.append(cell.coordinate)  # get core names if contain a 1
                cores.sort(key=lambda x: x[1:])
            values = np.array_split(values, ws.max_row)
            for row in values:
                rowcount.append(np.count_nonzero(row))  # EXCEL END
            self.cores = cores
            self.values = values
            self.rowcount = rowcount

    def info(self, text):
        self.label.setText(text)

    def export_images(self):
        self.progressBar.setMaximum(len(self.scene.centroid))
        self.export = Export(image=self.image, centroid=self.scene.centroid, cores=self.cores,
                             scale_index=self.scale_index,
                             core_diameter=self.core_diameter, output=self.output, name=self.name, lvl=self.lvl,
                             path=self.path)
        self.export.info.connect(self.info)
        self.export.countChanged.connect(self.progressBar.setValue)
        self.export.start()


class Export(QThread):
    info = pyqtSignal(str)
    countChanged = pyqtSignal(int)
    figures = pyqtSignal()

    def __init__(self, image, centroid, cores, scale_index, core_diameter, output, name, lvl, path):
        super().__init__()
        self.image = image
        self.centroid = centroid
        self.cores = cores
        self.scale_index = scale_index
        self.core_diameter = core_diameter
        self.output = output
        self.name = name
        self.lvl = lvl
        self.path = path

    def run(self):
        self.export_images(self.centroid, self.cores)

    def export_images(self, centroid, cores):
        infostr = []
        self.json_write()
        # lsize = size * self.scale_index
        scaledcent = [(y * self.scale_index, x * self.scale_index) for (x, y) in centroid]  # rotate xy for openslide
        scaledcent = [(int(x - (self.core_diameter / 2)), int(y - (self.core_diameter / 2))) for (x, y) in scaledcent]
        w_h = (self.core_diameter, self.core_diameter)
        self.lvl = 0
        for i in range(len(scaledcent)):
            infostr.append("Loading " + self.name + "_" + cores[i] + ".png")
            self.info.emit('\n'.join(infostr))
            core = self.image.read_region(location=scaledcent[i], level=self.lvl, size=w_h)
            mpimg.imsave(self.output + os.sep + self.name + "_" + cores[i] + ".png",
                         np.array(core))
            infostr.append("Saved " + self.name + "_" + cores[i] + ".png")
            self.info.emit('\n'.join(infostr))
            self.countChanged.emit(i + 1)

    def json_write(self):
        jsondata = {"path": self.path, "centroid": self.centroid, "cores": self.cores, "diameter": self.core_diameter,
                    "scale_index": self.scale_index}
        self.info.emit('Saving ' + self.output + os.sep + self.name + '_metadata.json')
        with open(self.output + os.sep + self.name + '_metadata.json', "w") as write_file:
            json.dump(jsondata, write_file)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MyWindow()

    # style sheet - https://github.com/mstuttgart/qdarkgraystyle
    app.setStyleSheet(qdarkgraystyle.load_stylesheet())

    sys.exit(app.exec_())
