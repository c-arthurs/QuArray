from PyQt5.uic import loadUi
from PyQt5.QtCore import Qt, QRectF, QThread, QObject, pyqtSignal, pyqtSlot
from PyQt5.QtGui import QColor, QFont, QImage, QPainter, QPixmap, QPen, QBrush
from PyQt5.QtWidgets import QGraphicsScene, QGraphicsItem, QFileDialog, QWidget, QGraphicsPixmapItem
from openslide import OpenSlide
import qimage2ndarray
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import json
import sys
from skimage.filters import threshold_otsu, threshold_li, threshold_mean, threshold_triangle, gaussian
from skimage.color import rgb2gray
from skimage.morphology import closing, square, remove_small_objects, label
from skimage.measure import regionprops
from skimage.transform import resize
from PIL import Image

if not hasattr(sys, "_MEIPASS"):
    sys._MEIPASS = '.'  # for running locally

# setup the Graphics scene to detect clicks
class GraphicsScene(QGraphicsScene):
    def __init__(self, parent=None):
        QGraphicsScene.__init__(self, parent)
        self.coords = []
        self.circles = []
        self.rectsandtext = []
        self.parent = MyWindow

    def reset(self):
        [self.circles[i].setVisible(False) for i in range(len(self.circles))]
        self.coords = []
        self.circles = []
        self.rectsandtext = []

    def elipse_adder(self, x, y):
        pen = QPen(QColor(69, 130, 201, 200))  # QColor(128, 128, 255, 128)
        pen.setWidthF(10)  # border width
        brush = QBrush(Qt.transparent)
        elipse = self.addEllipse(x - 25, y - 25, 50, 50, pen, brush)
        elipse.setAcceptDrops(True)
        elipse.setCursor(Qt.OpenHandCursor)
        elipse.setFlag(QGraphicsItem.ItemIsSelectable, True)
        elipse.setFlag(QGraphicsItem.ItemIsMovable, True)
        elipse.setFlag(QGraphicsItem.ItemIsFocusable, True)
        elipse.setFlag(QGraphicsItem.ItemSendsGeometryChanges, True)
        elipse.setAcceptHoverEvents(True)
        self.circles.append(elipse)

    def mouseDoubleClickEvent(self, event):
        x = event.scenePos().x()
        y = event.scenePos().y()
        btn = event.button()
        if btn == 1:  # left click
            self.coords.append((x, y))
            self.elipse_adder(x, y)

    def keyPressEvent(self, e): # hit space to remove point
        if e.key() == Qt.Key_Space:
            if len(self.circles) >= 1:
                self.circles[-1].setVisible(False)
                self.circles = self.circles[:-1]
                self.coords = self.coords[:-1]

    def sortCentroid(self, centroid):
        scent = sorted(centroid, key=lambda x: x[0])
        sortList = []
        a = 0
        comLength = 0
        for length in self.rowcount:
            comLength = comLength + length
            sortList.extend((sorted(scent[a:comLength], key=lambda k: [k[1]])))
            a = a + length
        return sortList

    def overlay_cores(self, core_diameter, scale_index, cores, autopilot=False):  # removed - centroid, image, cores
        if len(self.rectsandtext) >= 1:
            [i.setVisible(False) for i in self.rectsandtext]
            self.rectsandtext = []
        pen = QPen(QColor(69, 130, 201, 240))
        pen.setWidthF(6)  # border width
        brush = QBrush(QColor(215, 230, 248, 160))  # square fill
        if autopilot:
            self.centroid = self.coords
            self.centroid = self.sortCentroid(self.centroid)
            [self.elipse_adder(y, x) for (x, y) in self.centroid]
            self.coords = [(y, x) for (x, y) in self.centroid]
        else:
            self.centroid = [(y, x) for (x, y) in self.coords]
            self.centroid = [(self.centroid[i][0]+self.circles[i].scenePos().y(), self.centroid[i][1]+self.circles[i].scenePos().x()) for i in range(len(self.circles))]
        diameter = core_diameter / scale_index
        a = 0
        for y, x in self.centroid:
            try:
                rect = self.addRect((x - (diameter / 2)), (y - (diameter / 2)), diameter, diameter, pen, brush)
                text = self.addText(cores[a])  # label
                self.rectsandtext.append(rect)
                self.rectsandtext.append(text)
                text.setPos(x, y)
                text.setDefaultTextColor(QColor(35, 57, 82, 200))
                font = QFont()
                font.setPointSize(80)
                text.setFont(font)
                a = a + 1
            except IndexError as e:
                self.centroid.pop(a)
                print("index error", self.centroid[a])
                continue

    def save(self, output, name):
        # Get region of scene to capture from somewhere.
        area = self.sceneRect().size().toSize()
        image = QImage(area, QImage.Format_RGB888)
        painter = QPainter(image)
        self.render(painter, target=QRectF(image.rect()), source=self.sceneRect())
        painter.end()
        image.save(output + os.sep + name + "_overlay" + ".tiff")


class MyWindow(QWidget):
    def __init__(self):
        super(MyWindow, self).__init__()
        loadUi(sys._MEIPASS + os.sep + "scripts" + os.sep + "Cut_Application_thread_layout.ui", self)  # deployment
        self.setWindowTitle('QuArray')
        self.tabWidget.setStyleSheet("QTabWidget::pane {margin: 0px,0px,0px,0px; padding: 0px}")
        self.excel_layout = self.excel_btn.isChecked()
        self.excel_btn.toggled.connect(self.excel)
        self.load_ndpi.clicked.connect(lambda: self.loadndpi())
        self.load_excel.clicked.connect(lambda: self.read_excel())
        self.overlay.clicked.connect(lambda: self.overlaystart())
        self.export_2.clicked.connect(lambda: self.export_images())
        # self.export_again.clicked.connect(lambda: self.export_images(meta_only=True))
        self.current_image = None
        # threshold buttons
        self.gausianval = 0
        self.thresholdval = None
        self.otsu.clicked.connect(lambda: [self.threshold("otsu"), self.reset_sliders()])
        self.threshmean.clicked.connect(lambda: [self.threshold("mean"), self.reset_sliders()])
        self.threshtriangle.clicked.connect(lambda: [self.threshold("triangle"), self.reset_sliders()])
        self.threshli.clicked.connect(lambda: [self.threshold("li"), self.reset_sliders()])
        self.toggleorigional.clicked.connect(lambda: [self.threshold("origional"), self.reset_sliders()])
        self.gausslider.setMaximum(5)
        self.gausslider.setValue(0)
        self.gausslider.valueChanged.connect(lambda: self.gauslineEdit.setText(str(self.gausslider.value())))  # change
        self.gausslider.sliderReleased.connect(self.gaus)
        self.closingslider.setMaximum(50)
        self.closingslider.setValue(0)
        self.closingslider.valueChanged.connect(lambda: self.closelineEdit.setText(str(self.closingslider.value())))
        self.closingslider.sliderReleased.connect(self.closing)
        self.removesmallobjects.clicked.connect(self.removesmall)
        self.current_augments = {"threshold": False, "gausian": False, "closing": False, "overlay_applied": False,
                                 "manual_overlay": False}
        self.pathology = None
        self.excelpath = False

        self.init_scene()
        self.show()

    def overlaystart(self, autopilot=False, coords=None):
        """
        starts overlay
        :param autopilot: if the coords need to be asigned outside of the click function
        :param coords: the external coords to be applied in autopilot
        """
        try:
            self.core_diameter = int(self.diamiterLineEdit.text().strip())
        except:
            self.core_diameter = 6000
            self.diamiterLineEdit.setText('6000')
            self.info("core diamiter must be an integer - reset to 6000")
            pass
        if autopilot:
            self.scene.coords = coords
            self.scene.overlay_cores(self.core_diameter, self.scale_index, self.cores, autopilot=True)
        else:
            self.scene.overlay_cores(self.core_diameter, self.scale_index, self.cores)
        if self.overlaySave.isChecked():
            self.info(f"Overlay saved to - {self.output}")
            self.scene.save(self.output, self.name)
        self.current_augments["overlay_applied"] = True
        self.activate([self.export_2], action=True)

    def excel(self, x):
        self.excel_layout = x
        if self.excelpath:
            self.read_excel()

    def init_scene(self):
        self.scene = GraphicsScene(self)
        self.graphicsView.setScene(self.scene)
        self.pixmap = QGraphicsPixmapItem()
        self.scene.addItem(self.pixmap)

    def show_info(self, text):
        self.metadata.setText(text)

    def activate(self, names, action=True):
        for i in names:
            i.setEnabled(action)

    def loadndpi(self):
        self.init_scene()
        formats = '*.ndpi*;;*.svs*;;*.tif*;;*.scn*;;*.mrxs*;;*.tiff*;;*.svslide*;;*.bif*'
        self.path, _ = QFileDialog.getOpenFileName(parent=self, caption='Open file',
                                                   directory="/Users/callum/Desktop/", filter=formats)
        if self.path:
            self.output = os.path.splitext(self.path)[0] + '_split'
            if not os.path.exists(self.output):  # make output directory
                os.mkdir(self.output)

            self.name = os.path.split(self.output)[-1]
            self.nameLineEdit.setText(self.name)
            self.load_ndpi.setStyleSheet("background-color: rgb(0,90,0)")
            try:
                self.image = OpenSlide(self.path)
            except Exception as e:
                self.loadndpi()
            print(self.path + ' read to memory')
            print('    slide format = ' + str(OpenSlide.detect_format(self.path)))
            if str(OpenSlide.detect_format(self.path)) == "aperio":
                try:
                    print('    Magnification = ' + str(self.image.properties['openslide.objective-power']))  # TODO
                    print('    Date = ' + str(self.image.properties['aperio.Date']))
                    print('    dimensions = ' + str(self.image.dimensions))
                    print('    level_downsamples = ' + str(self.image.level_downsamples))
                except KeyError:
                    pass
            if str(OpenSlide.detect_format(self.path)) == "hamamatsu":
                try:
                    self.formatLineEdit.setText("Hamamatsu")
                    self.scanDateLineEdit.setText(str(self.image.properties['tiff.DateTime'][:10]))
                    self.dimensionsLineEdit.setText(str(self.image.dimensions))
                    self.magnificationLineEdit.setText(str(self.image.properties['hamamatsu.SourceLens']))
                    self.show_info(f"""Magnification = {str(self.image.properties['hamamatsu.SourceLens'])}
Date = {str(self.image.properties['tiff.DateTime'])}\ndimensions = {str(self.image.dimensions)}
level_downsamples = {str(self.image.level_downsamples)}""")
                    print('    Magnification = ' + str(self.image.properties['hamamatsu.SourceLens']))
                    print('    Date = ' + str(self.image.properties['tiff.DateTime']))
                    print('    dimensions = ' + str(self.image.dimensions))
                    print('    level_downsamples = ' + str(self.image.level_downsamples))
                    self.macro_image = self.image.associated_images['macro']
                except KeyError:
                    pass
            self.overview_level_width = 3000
            self.activate([self.nameLabel, self.nameLineEdit, self.formatLabel, self.formatLineEdit,
                           self.magnificationLabel, self.magnificationLineEdit, self.scanDateLabel,
                           self.scanDateLineEdit, self.dimensionsLabel, self.dimensionsLineEdit, self.overlayLevelLabel,
                           self.overlayLevelLineEdit, self.graphicsView, self.overlaySave, self.groupBox_2,
                           self.removesmallobjects])
            self.get_overview()
            if os.path.exists(os.path.splitext(self.path)[0] + '.xlsx'):
                self.excelpath = os.path.splitext(self.path)[0] + '.xlsx'
                self.read_excel()
            else:
                self.excelpath = False

    def get_overview(self):
        self.width_height = [
            (int(self.image.properties[f'openslide.level[{i}].width']),
             int(self.image.properties[f'openslide.level[{i}].height'])) for i
            in range(int(self.image.properties['openslide.level-count']))]
        width = [self.width_height[i][0] for i in range(len(self.width_height))]
        self.lvl = np.where(width == self.find_nearest(width, self.overview_level_width))[0][0]
        self.overlayLevelLineEdit.setText(str(self.lvl))
        self.scale_index = self.width_height[0][0] / self.width_height[self.lvl][0]
        self.overview = np.array(self.image.read_region(location=(0, 0), level=self.lvl,
                                                        size=self.width_height[self.lvl]))
        self.showimage(image=self.overview)

    def find_nearest(self, array, value):
        # https://stackoverflow.com/questions/2566412/find-nearest-value-in-numpy-array
        array = np.asarray(array)
        idx = (np.abs(array - value)).argmin()
        return array[idx]

    def showimage(self, image):
        self.current_image = image
        img = qimage2ndarray.array2qimage(image, normalize=True)
        img = QPixmap(img)
        self.pixmap.setPixmap(img)
        self.graphicsView.fitInView(self.graphicsView.sceneRect(), Qt.KeepAspectRatio)

    def reset_sliders(self):
        self.gausslider.setValue(0)
        self.closingslider.setValue(0)

    def threshold(self, threshold_name):
        if self.scene.circles:
            self.scene.reset()
        if self.current_augments['overlay_applied']:
            del self.scene.coords
            self.init_scene()
            self.read_excel()
            self.current_augments['overlay_applied'] = False
        if threshold_name == "origional":
            self.showimage(self.overview)
            self.thresholdval = None
            self.current_augments["threshold"] = False
            return
        self.thresholdval = threshold_name

        im = rgb2gray(self.overview)
        if threshold_name == "otsu":
            threshold = threshold_otsu(im)
        if threshold_name == "li":
            threshold = threshold_li(im)
        if threshold_name == "mean":
            threshold = threshold_mean(im)
        if threshold_name == "triangle":
            threshold = threshold_triangle(im)
        self.current_augments["threshold"] = threshold_name
        self.current_image = im < threshold
        self.showimage(self.current_image)

    def gaus(self):
        self.gauslineEdit.setText(str(self.gausslider.value()))
        if self.current_image.ndim > 2:
            filtered = gaussian(self.overview, sigma=self.gausslider.value())
        else:
            self.threshold(self.current_augments["threshold"])
            filtered = gaussian(self.current_image, sigma=self.gausslider.value())
        self.showimage(filtered)

    def closing(self):
        self.closelineEdit.setText(str(self.closingslider.value()))
        if self.current_augments["threshold"]:
            self.threshold(self.current_augments["threshold"])
            if self.closingslider.value() > 0:
                if self.current_augments["gausian"]:
                    self.current_image = gaussian(self.current_image, sigma=self.gausslider.value())
                closed = closing(self.current_image, square(self.closingslider.value()))
                closed = closed > 0
                self.showimage(closed)

    def removesmall(self):
        """
        the function name is misleading as this function now applies remove small objects - then
        overlays the cores and saves the image
        """
        if not self.current_augments['threshold']:
            return
        if self.current_augments['overlay_applied']:
            self.init_scene()
            self.read_excel()
        labeled_image = label(self.current_image)
        try:
            min = int(self.smallobs_text.text())
        except ValueError as e:
            self.smallobs_text.setText("6000")
            min = 6000
            self.info("remove small value must be an integer - reset to 6000")
        labeled_image = remove_small_objects(labeled_image, min_size=min)
        self.showimage(labeled_image)
        labels = regionprops(labeled_image)
        centroid = [r.centroid for r in labels]
        self.overlaystart(autopilot=True, coords=centroid)
        self.current_augments['overlay_applied'] = True

    def read_excel(self):
        if not self.excelpath:
            self.excelpath, _ = QFileDialog.getOpenFileName(parent=self, caption='Open file',
                                                            directory="/Users/callum/Desktop", filter="*.xlsx*")
        self.activate([self.numberOfCoresLabel, self.numberOfCoresLineEdit, self.diameterLabel, self.diamiterLineEdit,
                       self.overlay, self.progressBar, self.excel_btn, self.overlaySave, self.tabWidget])
        if self.excelpath:
            self.load_excel.setStyleSheet("background-color: rgb(0,90,0)")
            excelname = self.excelpath
            wb = load_workbook(excelname)
            ws = wb.worksheets[0]
            cores = []
            values = []
            rowcount = []
            self.rowcol = []
            if not self.excel_layout:
                for row in ws.iter_rows():
                    for cell in row:
                        values.append(cell.value)
                        if cell.value == 1:
                            cores.append((str(chr(int(cell.row) + 64))) + str(int(ord(cell.column_letter)) - 64))
            else:
                cores = []
                for col in ws.iter_cols():
                    for cell in col:
                        if cell.value == 1:
                            cores.append(cell.coordinate)  # get core names if contain a 1
                cores.sort(key=lambda x: x[1:])
                for row in ws.iter_rows():
                    for cell in row:
                        values.append(cell.value)
            values = np.array_split(values, ws.max_row)
            for row in values:
                rowcount.append(np.count_nonzero(row))  # EXCEL END
            self.numberOfCoresLineEdit.setText(str(len(cores)))
            self.cores = cores
            self.values = values
            self.rowcount = rowcount
            self.arrayshape = (ws.max_row, ws.max_column)
            self.scene.rowcount = rowcount
            # check for pathology file
            if len(wb.sheetnames) > 1:
                ws = wb.worksheets[1]  # pathology tab (hopefully)
                self.pathology = [ws[i].value for i in self.cores]

    def info(self, text):
        self.label.setText(text)

    def export_images(self, meta_only=False):
        if not self.scene.centroid:
            self.info("must overlay some selected core - double click on image")
            return
        self.progressBar.setMaximum(len(self.scene.centroid))
        self.activate([self.nameLabel, self.nameLineEdit, self.formatLabel, self.formatLineEdit,
                       self.magnificationLabel, self.magnificationLineEdit, self.scanDateLabel,
                       self.scanDateLineEdit, self.dimensionsLabel, self.dimensionsLineEdit, self.overlayLevelLabel,
                       self.overlayLevelLineEdit, self.graphicsView, self.numberOfCoresLabel,
                       self.numberOfCoresLineEdit, self.diameterLabel, self.diamiterLineEdit,
                       self.export_2, self.overlay, self.excel_btn, self.load_ndpi, self.load_excel, self.overlaySave,
                       self.tabWidget], action=False)
        try:
            resolution = int(self.resolution_edit.text())
            print("new resolution applied")
        except ValueError as E:
            resolution = 0

        self.export = Export(image=self.image, centroid=self.scene.centroid, cores=self.cores,
                             scale_index=self.scale_index,
                             core_diameter=self.core_diameter, output=self.output, name=self.name, lvl=self.lvl,
                            path=self.path, arrayshape=self.arrayshape, pathology=self.pathology, resolution=resolution, window=self,
                             meta_only=meta_only)

        self.thread = QThread()
        self.export.info.connect(self.info)
        self.export.done.connect(self.complete)
        self.export.countChanged.connect(self.progressBar.setValue)
        self.export.moveToThread(self.thread)
        self.thread.started.connect(self.export.run)
        self.thread.start()

    def complete(self):
        print('done')
        self.activate([self.nameLabel, self.nameLineEdit, self.formatLabel, self.formatLineEdit,
                       self.magnificationLabel, self.magnificationLineEdit, self.scanDateLabel,
                       self.scanDateLineEdit, self.dimensionsLabel, self.dimensionsLineEdit, self.overlayLevelLabel,
                       self.overlayLevelLineEdit, self.graphicsView, self.numberOfCoresLabel,
                       self.numberOfCoresLineEdit, self.diameterLabel, self.diamiterLineEdit,
                       self.export_2, self.overlay, self.excel_btn, self.load_ndpi, self.load_excel, self.overlaySave],
                      action=True)


class Export(QObject):
    info = pyqtSignal(str)
    countChanged = pyqtSignal(int)
    figures = pyqtSignal()
    done = pyqtSignal(bool)
    writemeta = pyqtSignal(bool)

    def __init__(self, image, centroid, cores, scale_index, core_diameter, output, name, lvl, path, arrayshape,
                 pathology, resolution, window, meta_only=False):
        super().__init__()
        self.image = image
        self.centroid = centroid
        self.cores = cores
        self.scale_index = scale_index
        self.core_diameter = core_diameter
        self.output = output
        self.name = name
        self.lvl = lvl
        self.path = path
        self.arrayshape = arrayshape
        self.pathology = pathology
        self.resolution = resolution
        self.meta_only = meta_only
        self.win = window

    @pyqtSlot()
    def run(self):
        if self.meta_only:
            self.json_write()
            self.wsifigure(higher_resolution=False, pathology=self.pathology)
        else:
            self.export_images(self.centroid, self.cores)

    def export_images(self, centroid, cores):
        infostr = []
        self.scaledcent = [(y * self.scale_index, x * self.scale_index) for (x, y) in centroid]  # rotate xy openslide
        self.scaledcent = [(int(x - (self.core_diameter / 2)), int(y - (self.core_diameter / 2))) for (x, y) in self.scaledcent]
        self.json_write()
        self.wsifigure(higher_resolution=False, pathology=self.pathology)
        w_h = (self.core_diameter, self.core_diameter)
        self.lvl = 0
        for i in range(len(self.scaledcent)):
            if not self.win.isVisible():
                print("need to kill here - main window closed")
                break
            infostr.append("Exporting " + self.name + "_" + cores[i] + ".png")
            print("Exporting " + self.name + "_" + cores[i] + ".png")
            self.info.emit('\n'.join(infostr))
            core = self.image.read_region(location=self.scaledcent[i], level=self.lvl, size=w_h)
            core.save(self.output + os.sep + self.name + "_" + cores[i] + ".png")
            infostr.append("Saved " + self.name + "_" + cores[i] + ".png")
            self.info.emit('\n'.join(infostr))
            self.countChanged.emit(i + 1)

        infostr.append("All files exported with JSON metadata")
        self.info.emit('\n'.join(infostr))
        print('\n'.join(infostr))
        self.done.emit(True)

    def json_write(self):
        jsondata = {"path": self.path, "coordinates": self.scaledcent, "cores": self.cores,
                    "diameter": self.core_diameter, "scale_index": self.scale_index, "lowlevel": int(self.lvl),
                    "arrayshape": self.arrayshape}
        self.info.emit('Saving ' + self.output + os.sep + self.name + '_metadata.json')
        with open(self.output + os.sep + self.name + '_metadata.json', "w") as write_file:
            json.dump(jsondata, write_file)

    def wsifigure(self, higher_resolution=False, pathology=None):
        """
        TODO: need to link this better to the actual script - rather than the meta file before its created
        todo: have removed button export_again for function later
        takes the metadata from the json
        makes a fig of the locations on the tissue array and saves it
        higher resolution = int start at 1 and move up to improve resolution - will slow code
        """
        higher_resolution = self.resolution

        def overly_path(im, overlay, pathology):
            if pathology == "N":
                colour = [0, 1, 0]  # green
            else:
                colour = [1, 0, 0]  # red
            im[overlay == 0] = colour
            return im

        jsonpath = self.output + os.sep + self.name + '_metadata.json'
        with open(jsonpath) as json_file:
            data = json.load(json_file)
            image = OpenSlide(data['path'])
            if higher_resolution:
                lvl = data['lowlevel'] - int(higher_resolution)
            else:
                lvl = data['lowlevel']
            scale_index = image.level_downsamples[lvl]
            diameter = int(data["diameter"] / scale_index)
            arrayshape = data['arrayshape']
            arrayshape = [(a * diameter) + 1 for a in arrayshape]
            arrayshape.append(3)
            figarray = np.ones(arrayshape)
            cindex = [coordinate_from_string(i) for i in data['cores']]  # returns ('A',4)
            cindex = [(b, column_index_from_string(a)) for (a, b) in cindex]  # returns index tuples for each core
            maskcoord = [[y * diameter - diameter if y > 1 else y for y in x] for x in cindex]

            if pathology:
                overlay = np.load(sys._MEIPASS + os.sep + "scripts" + os.sep + "outline.npy")
                overlay = resize(overlay, (diameter, diameter))

            for i in range(len(data["coordinates"])):  # iterate through coordinates pulling out images from wsi
                im = image.read_region(location=data["coordinates"][i], level=lvl, size=(diameter, diameter))
                im = np.array(im)[:, :, :3]
                if pathology:
                    try:
                        im = overly_path(im, overlay, pathology[i])
                    except:
                        continue
                figarray[maskcoord[i][0]:maskcoord[i][0] + diameter, maskcoord[i][1]:maskcoord[i][1] + diameter, :] = im
            figarray[figarray == [1, 1, 1]] = 255  # make background white
            savepath = self.output + os.sep + self.name + '_layoutfig.tiff'
            Image.fromarray(figarray.astype(np.uint8)).save(savepath)
